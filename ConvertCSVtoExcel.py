import pandas as pd
from openpyxl import load_workbook
import random as rand
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
import glob, os


def CSVtoExcel(CSV):
    Original_CSV = pd.read_csv(CSV)
    New_Excel = pd.ExcelWriter('C:/Users/adiaz/Desktop/ZoomAttendenceSheet/CSVFile/AttendanceReport_' + str(
        rand.randrange(1000, 9999)) + '.xlsx')
    Original_CSV.to_excel(New_Excel, index=False)

    New_Excel.save()

    wb = load_workbook(New_Excel)
    ws = wb['Sheet1']

    count = 3
    for col_cells in ws.iter_cols(min_row=4, min_col=7, max_col=7):
        for cell in col_cells:
            count += 1
            cell.value = '=E' + str(count) + '/$F$2'
            ws['G' + str(count)].style = 'Percent'
            ws['G' + str(count)].font = Font(bold=True)

    redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
    greenFill = PatternFill(start_color='00ff00', end_color='00ff00', fill_type='solid')

    ws.conditional_formatting.add('G4:' + 'G' + str(len(ws['G'])),
                                  CellIsRule('lessThan', formula=['0.7'], stopIfTrue=True, fill=redFill))
    ws.conditional_formatting.add('G4:' + 'G' + str(len(ws['G'])),
                                  CellIsRule('greaterThan', formula=['0.7'], stopIfTrue=True, fill=greenFill))

    wb.save(New_Excel)

    return max(glob.iglob('CSVFile/*.xlsx'), key=os.path.getctime)
